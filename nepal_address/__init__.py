# -*- coding: utf-8 -*-
from . import models

import os
import base64
import tempfile
import openpyxl
from odoo import api, SUPERUSER_ID

def post_init_hook(env):
    print('calling post init hook')
    module_path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(module_path, 'data', 'res.province (res.province)(1).xlsx')

    workbook = openpyxl.load_workbook(file_path)
    
    # Get Nepal country
    country = env['res.country'].search([('code', '=', 'NP')], limit=1)
    if not country:
        print("Nepal country not found!")
        return

    # 1. Load Provinces
    sheet = workbook['Province']
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        name = row[0]
        if name:
            env['res.province'].create({
                'name': name,
                'country_id': country.id,
            })

    # 2. Load Districts
    sheet = workbook['District']
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        name, province_name = row[0], row[1]
        province = env['res.province'].search([('name', '=', province_name)], limit=1)
        if name and province:
            env['res.district'].create({
                'name': name,
                'province_id': province.id,
                'country_id': country.id,
            })

    # 3. Load Municipalities
    sheet = workbook['Municipality']
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        name, district_name = row[0], row[1]
        district = env['res.district'].search([('name', '=', district_name)], limit=1)
        if name and district:
            env['res.municipality'].create({
                'name': name,
                'district_id': district.id,
                'country_id': country.id,
            })

    # 4. Load Wards
    sheet = workbook['Ward']
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        name, municipality_name = row[0], row[1]
        municipality = env['res.municipality'].search([('name', '=', municipality_name)], limit=1)
        if name and municipality:
            env['res.ward'].create({
                'name': str(name),  # Ensure ward number is stored as string
                'municipality_id': municipality.id,
                'district_id': municipality.district_id.id,
                'country_id': country.id,
            })